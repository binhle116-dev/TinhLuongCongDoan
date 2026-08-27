"""Nap du lieu gia/anh xa THAT vao ServiceCategory/PriceGroup/PriceCard/
RouteGroupMapping/ServiceMapping, can cu vao 2 nguon du lieu that:

1. File chu hang thang (mac dinh: file thang 07/2026) - sheet
   "Don gia XD 2026" (bang gia 10 nhom x 49 dich vu) va sheet "Tuyen"
   (ma tuyen -> nhom gia, da xac minh khop 100% - 117/117 - voi
   ROUTE_PO_CODE that trong du lieu tho da import).
2. Cac to hop SERVICE_NAME_PAYROLL/TYPE_CODE_PAYROLL/AREA_CODE/can nang
   THAT xuat hien trong RawDailyProduction da import (ngay 26/08/2026,
   6706 dong) - dung de xay quy tac ServiceMapping co can cu, KHONG
   doan cho cac truong hop con mo (xem PROJECT_CONTEXT.md Section 6):
   "C-Bao Phat", "Goi nho thuong", cac bien the "KT1 ...- B/C" chua ro
   co phai KT1 ABC hay khong, va "KT1 Hoa toc Hen gio" (ket hop 2 thuoc
   tinh, khong co danh muc rieng) - CO CHU DICH de trong, se hien trong
   /bao-cao/chua-anh-xa/ cho Admin/TCHC xu ly.

Idempotent: chay lai nhieu lan an toan (dung update_or_create).
"""
from pathlib import Path

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from django.utils.text import slugify

from phat.models import PriceCard, PriceGroup, RouteGroupMapping, ServiceCategory, ServiceMapping

DEFAULT_MASTER_FILE = (
    Path(settings.BASE_DIR).parent / "CÔNG PHÁT THÁNG 07.2026 - ttvh.xlsx"
)

WEIGHT_2KG = 2000.0

# (service_name_payroll, type_code_payroll, area_code, weight_max_gram or None, weight_min_gram or None, category_name)
# weight_max=2000 => tier "<=2kg"; weight_min=2000.01 => tier ">2kg" (khong trung nhau).
SERVICE_MAPPING_RULES = [
    # R-Bưu phẩm bảo đảm (khong phan biet COD, chi co 1 loai TYPE thay doi)
    ("R-Bưu phẩm bảo đảm", None, "NT", None, None, "BPBD nội tỉnh"),
    ("R-Bưu phẩm bảo đảm", None, "LT", None, None, "BPBD liên tỉnh"),
    ("R-Bưu phẩm bảo đảm", None, "QT", None, None, "Bưu phẩm đảm bảo quốc tế"),
    # Báo phát (R/E) - khong phan biet COD/can nang
    ("R-Báo Phát", None, "NT", None, None, "Báo phát R Nội Tỉnh"),
    ("R-Báo Phát", None, "LT", None, None, "Báo phát R Liên Tỉnh"),
    ("E-Báo Phát", None, "NT", None, None, "Báo phát E Nội Tỉnh"),
    ("E-Báo Phát", None, "LT", None, None, "Báo phát E Liên Tỉnh"),
    # Hoa toc E (E-Hoa toc, khac voi "EMS(tru...)")
    ("E-Hỏa tốc", None, "NT", None, None, "EMS hỏa tốc nội tỉnh"),
    ("E-Hỏa tốc", None, "LT", None, None, "EMS hỏa tốc liên tỉnh"),
    # KT1 thuong (type=KT1)
    ("KT1", "KT1", "NT", None, None, "KT1 Nội Tỉnh"),
    ("KT1", "KT1", "LT", None, None, "KT1 liên tỉnh"),
    # KT1 hoa toc dung ten (khong co hau to ABC/-B/-C)
    ("KT1 Hỏa tốc", "KT1_HT", "NT", None, None, "KT1 Hỏa Tốc Nội tỉnh"),
    ("KT1 Hỏa tốc", "KT1_HT", "LT", None, None, "KT1 hỏa tốc liên tỉnh"),
    # APP-Epacket
    ("L-AppEpacket", None, "QT", None, None, "APP-Epacket"),
    # C-Bưu kiện: KCOD (khong COD) - phan theo can nang
    ("C-Bưu kiện", "DV_T_KCOD", "NT", WEIGHT_2KG, None, "BK nội tỉnh <= 2Kg"),
    ("C-Bưu kiện", "DV_T_KCOD", "NT", None, WEIGHT_2KG + 0.01, "Bưu kiện nội tỉnh > 2Kg"),
    ("C-Bưu kiện", "DV_T_KCOD", "LT", WEIGHT_2KG, None, "BK liên tỉnh <= 2Kg"),
    ("C-Bưu kiện", "DV_T_KCOD", "LT", None, WEIGHT_2KG + 0.01, "Bưu kiện liên tỉnh > 2Kg"),
    ("C-Bưu kiện", "DV_T_KCOD_GTGT", "NT", WEIGHT_2KG, None, "BK nội tỉnh <= 2Kg"),
    ("C-Bưu kiện", "DV_T_KCOD_GTGT", "NT", None, WEIGHT_2KG + 0.01, "Bưu kiện nội tỉnh > 2Kg"),
    ("C-Bưu kiện", "DV_T_KCOD_GTGT", "LT", WEIGHT_2KG, None, "BK liên tỉnh <= 2Kg"),
    ("C-Bưu kiện", "DV_T_KCOD_GTGT", "LT", None, WEIGHT_2KG + 0.01, "Bưu kiện liên tỉnh > 2Kg"),
    # C-Bưu kiện: co COD - phan theo can nang
    ("C-Bưu kiện", "DV_T_COD", "NT", WEIGHT_2KG, None, "Bưu kiện COD nội tỉnh <= 2kg"),
    ("C-Bưu kiện", "DV_T_COD", "NT", None, WEIGHT_2KG + 0.01, "Bưu kiện COD nội tỉnh > 2Kg"),
    ("C-Bưu kiện", "DV_T_COD", "LT", WEIGHT_2KG, None, "Bưu kiện COD liên tỉnh <= 2Kg"),
    ("C-Bưu kiện", "DV_T_COD", "LT", None, WEIGHT_2KG + 0.01, "Bưu kiện COD liên tỉnh > 2Kg"),
    ("C-Bưu kiện", "DV_T_COD_GTGT", "NT", WEIGHT_2KG, None, "Bưu kiện COD nội tỉnh <= 2kg"),
    ("C-Bưu kiện", "DV_T_COD_GTGT", "NT", None, WEIGHT_2KG + 0.01, "Bưu kiện COD nội tỉnh > 2Kg"),
    ("C-Bưu kiện", "DV_T_COD_GTGT", "LT", WEIGHT_2KG, None, "Bưu kiện COD liên tỉnh <= 2Kg"),
    ("C-Bưu kiện", "DV_T_COD_GTGT", "LT", None, WEIGHT_2KG + 0.01, "Bưu kiện COD liên tỉnh > 2Kg"),
    # E-EMS(tru Bao phat va Hoa toc): KCOD - phan theo can nang
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_KCOD", "NT", WEIGHT_2KG, None, "EMS nội tỉnh <=2kg"),
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_KCOD", "NT", None, WEIGHT_2KG + 0.01, "EMS nội tỉnh >2kg"),
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_KCOD", "LT", WEIGHT_2KG, None, "EMS liên tỉnh <=2kg"),
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_KCOD", "LT", None, WEIGHT_2KG + 0.01, "EMS liên tỉnh >2kg"),
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_KCOD_GTGT", "NT", WEIGHT_2KG, None, "EMS nội tỉnh <=2kg"),
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_KCOD_GTGT", "NT", None, WEIGHT_2KG + 0.01, "EMS nội tỉnh >2kg"),
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_KCOD_GTGT", "LT", WEIGHT_2KG, None, "EMS liên tỉnh <=2kg"),
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_KCOD_GTGT", "LT", None, WEIGHT_2KG + 0.01, "EMS liên tỉnh >2kg"),
    # E-EMS(tru...): co COD - phan theo can nang
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_COD", "NT", WEIGHT_2KG, None, "EMS COD nội tỉnh <=2kg"),
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_COD", "NT", None, WEIGHT_2KG + 0.01, "EMS COD nội tỉnh >2kg"),
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_COD", "LT", WEIGHT_2KG, None, "EMS COD liên tỉnh <=2kg"),
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_COD", "LT", None, WEIGHT_2KG + 0.01, "EMS COD liên tỉnh >2kg"),
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_COD_GTGT", "NT", WEIGHT_2KG, None, "EMS COD nội tỉnh <=2kg"),
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_COD_GTGT", "NT", None, WEIGHT_2KG + 0.01, "EMS COD nội tỉnh >2kg"),
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_COD_GTGT", "LT", WEIGHT_2KG, None, "EMS COD liên tỉnh <=2kg"),
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", "DV_N_COD_GTGT", "LT", None, WEIGHT_2KG + 0.01, "EMS COD liên tỉnh >2kg"),
    # E-EMS(tru...) quoc te - khong phan COD/can nang (chua thay du lieu that phan biet)
    ("E-EMS(trừ E-Báo phát và E-Hỏa tốc)", None, "QT", None, None, "EMS Quốc tế"),
]

# Cac to hop THAT da xac nhan la con mo, CO CHU DICH khong tao quy tac -
# se hien trong /bao-cao/chua-anh-xa/. Ghi lai o day de tra cuu, khong
# dung de tao du lieu.
KNOWN_AMBIGUOUS_COMBINATIONS = [
    ("C-Báo Phát", "DV_T_KCOD", "LT"),
    ("Gói nhỏ thường", "BPBDQT", "QT"),
    ("KT1 Hỏa tốc - C", "KT1_HT", "NT"),
    ("KT1 C", "KT1_HT", "NT"),
    ("KT1 C", "KT1_HT", "LT"),
    ("KT1 Hỏa tốc - C", "KT1_HT", "LT"),
    ("KT1 Hỏa tốc - B", "KT1_HT", "LT"),
    ("KT1 B", "KT1_HT", "LT"),
    ("KT1 Hỏa tốc Hẹn giờ", "KT1_HT", "NT"),
]


class Command(BaseCommand):
    help = (
        "Nap du lieu gia/anh xa THAT (ServiceCategory, PriceGroup, PriceCard, "
        "RouteGroupMapping, ServiceMapping) tu file chu hang thang. Idempotent."
    )

    def add_arguments(self, parser):
        parser.add_argument("--file", dest="file", default=str(DEFAULT_MASTER_FILE))

    def handle(self, *args, **options):
        master_path = Path(options["file"])
        if not master_path.exists():
            self.stderr.write(self.style.ERROR(f"Khong tim thay file: {master_path}"))
            return

        wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)

        n_cat, n_group, n_price = self._seed_price_table(wb)
        self.stdout.write(self.style.SUCCESS(
            f"Da nap: {n_cat} loai dich vu, {n_group} nhom gia, {n_price} dong bang gia."
        ))

        n_route = self._seed_route_groups(wb)
        self.stdout.write(self.style.SUCCESS(f"Da nap: {n_route} anh xa Tuyen -> Nhom gia."))

        n_mapping = self._seed_service_mapping()
        self.stdout.write(self.style.SUCCESS(f"Da nap: {n_mapping} quy tac anh xa dich vu."))

        self.stdout.write(
            f"Con {len(KNOWN_AMBIGUOUS_COMBINATIONS)} to hop CO CHU DICH chua tao quy tac "
            "(xem /bao-cao/chua-anh-xa/ va PROJECT_CONTEXT.md Section 6)."
        )

    def _seed_price_table(self, wb):
        ws = wb["Đơn giá XD 2026"]
        rows = list(ws.iter_rows(values_only=True))
        header = [c for c in rows[0] if c]  # bo cot rong cuoi sheet
        category_names = header[2:]  # bo 'Tên nhóm', 'Mã nhóm'

        # QUAN TRONG: slugify() bo ky tu "<=" / ">" nen "EMS noi tinh <=2kg"
        # va "EMS noi tinh >2kg" se trung ma neu chi dung slugify(name) -
        # phai ghep them chi so cot de dam bao ma khong bao gio trung.
        categories = {}
        for idx, name in enumerate(category_names):
            code = f"svc-{idx:02d}-{slugify(name, allow_unicode=False)[:40]}"
            obj, _ = ServiceCategory.objects.update_or_create(
                code=code, defaults={"name": name}
            )
            categories[name] = obj

        # Sheet co bang bi lap lai (dong 13-23) + 1 bang "Don gia 2025"/
        # "Don gia 90%" khac phia duoi (dong 26+) - chi lay dong bat dau
        # bang "NHOM <so>" that su, bo qua moi thu khac de tranh doc
        # nham gia tri sai.
        n_group = 0
        n_price = 0
        seen_groups = set()
        for r in rows[1:]:
            label = str(r[0] or "").strip()
            if not label.upper().startswith("NHÓM") or r[1] is None:
                continue
            try:
                group_code = int(r[1])
            except (TypeError, ValueError):
                continue
            if group_code in seen_groups:
                continue
            seen_groups.add(group_code)
            group, created = PriceGroup.objects.update_or_create(
                code=group_code, defaults={"name": str(r[0])}
            )
            n_group += 1 if created else 0
            for idx, name in enumerate(category_names, start=2):
                price = r[idx]
                if price is None:
                    continue
                PriceCard.objects.update_or_create(
                    service_category=categories[name],
                    price_group=group,
                    effective_from=None,
                    defaults={"unit_price": round(float(price), 2)},
                )
                n_price += 1

        return len(categories), PriceGroup.objects.count(), n_price

    def _seed_route_groups(self, wb):
        ws = wb["Tuyến"]
        rows = list(ws.iter_rows(values_only=True))
        n = 0
        for r in rows[3:]:
            if not r[0] or r[2] is None:
                continue
            route_code = str(r[0]).strip()
            if route_code.endswith(".0"):
                route_code = route_code[:-2]
            group_code = int(r[2])
            try:
                group = PriceGroup.objects.get(code=group_code)
            except PriceGroup.DoesNotExist:
                continue
            RouteGroupMapping.objects.update_or_create(
                route_code=route_code, effective_from=None, defaults={"price_group": group}
            )
            n += 1
        return n

    def _seed_service_mapping(self):
        n = 0
        for priority, (name, type_code, area, weight_max, weight_min, category_name) in enumerate(
            SERVICE_MAPPING_RULES, start=100
        ):
            try:
                category = ServiceCategory.objects.get(name=category_name)
            except ServiceCategory.DoesNotExist:
                self.stderr.write(self.style.WARNING(
                    f"Bo qua quy tac - khong tim thay loai dich vu '{category_name}'"
                ))
                continue
            ServiceMapping.objects.update_or_create(
                service_name_payroll=name,
                type_code_payroll=type_code or "",
                area_code=area,
                weight_min_gram=weight_min,
                weight_max_gram=weight_max,
                defaults={
                    "priority": priority,
                    "service_category": category,
                    "is_active": True,
                    "note": "Nap tu seed_pricing_data (can cu du lieu that + bang Don gia XD 2026)",
                },
            )
            n += 1
        return n
