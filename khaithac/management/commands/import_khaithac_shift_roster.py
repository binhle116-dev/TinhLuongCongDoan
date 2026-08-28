"""Import Bang cham cong Khai thac that ("BCC he so <nam> (LT).xlsx", 1
sheet/thang, vd 'T7.26' cho thang 7/2026) vao KhaiThacShiftAssignment.

Cau truc that cua 1 sheet (xac minh thu cong voi T7.26, khop chinh xac
dong TONG cua file - 302.6): cot A=NGAY (chi co gia tri o dong dau moi
ngay), B=CONG VIEC (TRUONG CA/KTV/GSS...), C=TEN, D=HE SO HDC, E=CA
(1/2/3, de trong voi GSS). Dong TEN trong = cho trong (chua xep nguoi),
bo qua. Dong cuoi sheet la dong TONG (cot A la text) - dung lai truoc do.

Ten tren bang cham cong ghi hoa/thuong binh thuong, con Employee.full_name
trong he thong luu HOA TOAN BO - so khop qua .upper() bang Python (KHONG
dung Employee.objects.filter(full_name__iexact=...) vi SQLite UPPER() mac
dinh khong xu ly dung chu co dau tieng Viet).

Idempotent theo (post_office, year, month): xoa het dong cu truoc khi ghi
lai.

Cach chay:
    python manage.py import_khaithac_shift_roster --thang 2026-07
"""
from __future__ import annotations

import datetime as dt
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.models import Employee, PostOffice
from khaithac.models import KhaiThacShiftAssignment

POST_OFFICE_CODE = "530100"
DEFAULT_FILE_DIR = Path(
    r"D:\ONEDRIVE\Trung tam Van hanh - BDTP Hue 05.2025\2026\San luong tinh luong cac cong doan\Khai thac"
)

CA_MAP = {1: "CA1", 2: "CA2", 3: "CA3"}


def _find_default_file() -> Path:
    matches = sorted(DEFAULT_FILE_DIR.glob("BCC*.xlsx"))
    if not matches:
        raise CommandError(f"Khong tim thay file 'BCC*.xlsx' trong {DEFAULT_FILE_DIR}")
    return matches[-1]


def _sheet_name_candidates(month: int, year: int) -> list[str]:
    yy = year % 100
    return [f"T{month}.{yy}", f"T{month},{yy}"]  # T2,26 co dau phay thay vi cham trong file that


class Command(BaseCommand):
    help = "Import Bang cham cong Khai thac (he so ca that) tu file Excel cua don vi."

    def add_arguments(self, parser):
        parser.add_argument("--thang", required=True, help="Thang can import, VD 2026-07")
        parser.add_argument("--file", required=False, help="Duong dan file .xlsx (mac dinh: tim BCC*.xlsx moi nhat)")

    def handle(self, *args, **options):
        year, month = (int(x) for x in options["thang"].split("-"))
        file_path = Path(options["file"]) if options.get("file") else _find_default_file()
        if not file_path.exists():
            raise CommandError(f"Khong tim thay file {file_path}")

        try:
            post_office = PostOffice.objects.get(code=POST_OFFICE_CODE)
        except PostOffice.DoesNotExist:
            raise CommandError(f"Khong tim thay PostOffice code={POST_OFFICE_CODE}")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_name = next((s for s in _sheet_name_candidates(month, year) if s in wb.sheetnames), None)
        if sheet_name is None:
            raise CommandError(
                f"Khong tim thay sheet cho thang {month}/{year} trong {file_path.name} "
                f"(da thu: {_sheet_name_candidates(month, year)}, sheet co san: {wb.sheetnames})"
            )
        ws = wb[sheet_name]

        employees_by_upper: dict[str, Employee] = {}
        for emp in Employee.objects.all():
            employees_by_upper.setdefault(emp.full_name.strip().upper(), emp)

        cur_day: int | None = None
        parsed_rows = []
        for r in range(3, ws.max_row + 1):
            day_cell = ws.cell(row=r, column=1).value
            if isinstance(day_cell, int):
                cur_day = day_cell
            elif isinstance(day_cell, str):
                break  # dong TONG hoac header lap lai - het du lieu that

            cong_viec = ws.cell(row=r, column=2).value
            ten = ws.cell(row=r, column=3).value
            he_so = ws.cell(row=r, column=4).value
            ca_so = ws.cell(row=r, column=5).value

            if not ten or he_so is None or cur_day is None:
                continue  # cho trong (chua xep nguoi)

            parsed_rows.append(
                {
                    "work_date": dt.date(year, month, cur_day),
                    "cong_viec": (cong_viec or "").strip(),
                    "raw_name": ten.strip(),
                    "he_so": he_so,
                    "ca": CA_MAP.get(ca_so, ""),
                }
            )

        unmatched_names: dict[str, int] = {}
        to_create = []
        for row in parsed_rows:
            employee = employees_by_upper.get(row["raw_name"].upper())
            if employee is None:
                unmatched_names[row["raw_name"]] = unmatched_names.get(row["raw_name"], 0) + 1
            to_create.append(
                KhaiThacShiftAssignment(
                    employee=employee, raw_name=row["raw_name"], work_date=row["work_date"],
                    cong_viec=row["cong_viec"], ca=row["ca"], he_so=row["he_so"],
                )
            )

        with transaction.atomic():
            KhaiThacShiftAssignment.objects.filter(
                work_date__year=year, work_date__month=month
            ).delete()
            KhaiThacShiftAssignment.objects.bulk_create(to_create)

        total_he_so = sum(r["he_so"] for r in parsed_rows)
        self.stdout.write(self.style.SUCCESS(
            f"Da import {len(to_create)} dong tu sheet '{sheet_name}' ({file_path.name}). "
            f"Tong he so: {total_he_so}."
        ))
        if unmatched_names:
            self.stdout.write(self.style.WARNING(
                f"Chua khop duoc {len(unmatched_names)} ten voi nhan vien nao trong he thong "
                f"(van luu lai, khong bi mat khoi tong he so): {unmatched_names}"
            ))
